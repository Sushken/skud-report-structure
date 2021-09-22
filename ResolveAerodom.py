import openpyxl
import xlrd
import time
import os
from PyQt5 import QtCore
from datetime import datetime


class Resolve(QtCore.QThread):
    percentageChanged = QtCore.pyqtSignal(int)
    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, nameOfInFile, count):
        super().__init__()
        self.nameOfInFile = nameOfInFile
        self.count = count

    def run(self):
        start_time = datetime.now()
        self.workbook = xlrd.open_workbook(self.nameOfInFile)
        self.worksheet = self.workbook.sheet_by_index(0)
        dataCompany = []
        dataStructure = []
        dataPosition = []
        dataFIO = []
        dataGender = []
        dataFirstIn = []
        dataLastOut = []
        row = 7
        while row != self.worksheet.nrows:
            dataCompany.append(self.worksheet.cell_value(row, 0)) #2,5,7,9,11,14
            dataStructure.append(self.worksheet.cell_value(row, 2))
            dataPosition.append(self.worksheet.cell_value(row, 5))
            dataFIO.append(self.worksheet.cell_value(row, 7))
            dataGender.append(self.worksheet.cell_value(row, 9))
            dataFirstIn.append(self.worksheet.cell_value(row, 11))
            dataLastOut.append(self.worksheet.cell_value(row, 14))
            row += 1

        for j in range(25):
            time.sleep(0.2)
            self.count += 1
            self.percentageChanged.emit(self.count)

        tmp_dataCompany = []
        tmp_dataStructure = []
        tmp_dataPosition = []
        tmp_dataFIO = []
        tmp_dataGender = []
        tmp_dataFirstIn = []
        tmp_dataLastOut = []
        for member in range(len(dataCompany)):
            if dataCompany[member] != '' and dataCompany[member] != 'Компания':
                tmp_dataCompany.append(dataCompany[member])
                tmp_dataStructure.append(dataStructure[member])
                tmp_dataPosition.append(dataPosition[member])
                tmp_dataFIO.append(dataFIO[member])
                tmp_dataGender.append(dataGender[member])
                tmp_dataFirstIn.append(dataFirstIn[member])
                tmp_dataLastOut.append(dataLastOut[member])

        for j in range(25):
            self.count += 1
            self.percentageChanged.emit(self.count)

        self.filepath = os.path.expanduser("~/Desktop/Новый Аэродом.xlsx")
        wbOut = openpyxl.Workbook()
        sheetOut = wbOut['Sheet']

        i = 0
        while i != len(tmp_dataCompany):
            sheetOut.cell(row=i + 1, column=1).value = tmp_dataCompany[i]
            sheetOut.cell(row=i + 1, column=2).value = tmp_dataStructure[i]
            sheetOut.cell(row=i + 1, column=3).value = tmp_dataPosition[i]
            sheetOut.cell(row=i + 1, column=4).value = tmp_dataFIO[i]
            sheetOut.cell(row=i + 1, column=5).value = tmp_dataGender[i]
            sheetOut.cell(row=i + 1, column=6).value = tmp_dataFirstIn[i]
            sheetOut.cell(row=i + 1, column=7).value = tmp_dataLastOut[i]
            i += 1

        for j in range(25):
            self.count += 1
            self.percentageChanged.emit(self.count)

        wbOut.save(self.filepath)

        for j in range(25):
            self.count += 1
            self.percentageChanged.emit(self.count)
        end_time = datetime.now()
        self.duration = end_time - start_time
        self.indicator = True
        self.indicator_of_end_work.emit(self.indicator)
