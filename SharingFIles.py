import openpyxl
import xlrd
import xlsxwriter
import os
from PyQt5 import QtCore
from datetime import datetime


class CreateFiles:
    # percentageChanged = QtCore.pyqtSignal(int)
    # indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, company, fileName):
        super().__init__()
        self.company = company
        self.nameOfFiles = fileName
        print("Start this prog!")
        self.db = "БД.xlsx"
        self.oldDataDepartments, self.oldDataPath = self.openFile(self.db)
        self.dataDepartments = []
        self.dataPath = []
        for i in self.oldDataDepartments:
            if i != '':
                self.dataDepartments.append(i)
        for i in self.oldDataPath:
            if i != '':
                self.dataPath.append(i)
        self.nameOfFilesXlsx = self.nameOfFiles + ".xlsx"

    def run(self):
        for count in range(len(self.dataPath)):
            filepath = os.path.expanduser(self.dataPath[count] + self.nameOfFilesXlsx)
            wbOut = openpyxl.Workbook()
            wbOut.create_sheet("Лист1", index=0)
            wbOut.save(filepath)
            wbOut.close()

    def openFile(self, file):
        workbook = xlrd.open_workbook(file)
        worksheet = workbook.sheet_by_index(0)
        dataDepartment = []
        dataPath = []
        row = 3
        # select_company = input()
        if self.company == "ГК Новотранс":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 0))
                dataPath.append(worksheet.cell_value(row, 1))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "РК Новотранс":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 4))
                dataPath.append(worksheet.cell_value(row, 5))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Новотранс Актив":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 2))
                dataPath.append(worksheet.cell_value(row, 3))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "ХК Новотранс":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 6))
                dataPath.append(worksheet.cell_value(row, 7))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Арго":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 8))
                dataPath.append(worksheet.cell_value(row, 9))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "УК Новотранс":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 10))
                dataPath.append(worksheet.cell_value(row, 11))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (НВТА)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 12))
                dataPath.append(worksheet.cell_value(row, 13))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (БТП)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 14))
                dataPath.append(worksheet.cell_value(row, 15))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (КУЛ)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 16))
                dataPath.append(worksheet.cell_value(row, 17))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (СК)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 18))
                dataPath.append(worksheet.cell_value(row, 19))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (БВРЗ)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 20))
                dataPath.append(worksheet.cell_value(row, 21))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (УК)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 22))
                dataPath.append(worksheet.cell_value(row, 23))
                row += 1
            return dataDepartment, dataPath
        elif self.company == "Питер (Строй)":
            while row != worksheet.nrows:
                dataDepartment.append(worksheet.cell_value(row, 24))
                dataPath.append(worksheet.cell_value(row, 25))
                row += 1
            return dataDepartment, dataPath


class WorkWithFile(QtCore.QThread):

    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, path, company, nameOfFiles):
        super().__init__()
        self.workFile = path
        self.company = company
        self.nameOfFiles = nameOfFiles
        print("Start work with files!")
        self.header = [
            'Компания',
            'Отдел',
            'Должность',
            'Пол',
            'Фамилия',
            'Имя',
            'Отчество',
            'Первый приход в любой из офисов',
            'Последний уход из любого офиса',
            'Продолжительность рабочего дня',
            'Обеденный перерыв',
        ]
        # self.run()

    def run(self):
        self.fileCreator = CreateFiles(self.company, self.nameOfFiles)
        self.fileCreator.run()
        self.work()

    def work(self):
        data, countRows = self.readFile(self.workFile)
        print("DATA = ", data)
        start = datetime.now()
        new_data = []
        for i in range(0, len(data) - 10, 11):
            new_data.append(data[i:i+11])
        print(new_data)
        print(self.fileCreator.dataPath[0] + self.fileCreator.nameOfFilesXlsx)

        for dbCount in range(len(self.fileCreator.dataDepartments)):
            i = 0
            ifNoEntry = True
            workbook = xlsxwriter.Workbook(self.fileCreator.dataPath[dbCount] + self.fileCreator.nameOfFilesXlsx)
            worksheet = workbook.add_worksheet("Лист1")
            date_format = workbook.add_format({'num_format': '[hh]:mm:ss'})
            set_row = 1

            while i != len(new_data):
                if new_data[i][1] == self.fileCreator.dataDepartments[dbCount]:
                    worksheet.write_row(set_row, 0, new_data[i])
                    set_row += 1
                    ifNoEntry = False
                i += 1
            worksheet.set_column('A:K', width=30)
            worksheet.set_column('K:K', None, None, {'hidden': 1})
            worksheet.set_column('J:J', width=20, cell_format=date_format)
            worksheet.set_row(0, height=50)
            first_row_format = workbook.add_format()
            first_row_format.set_bg_color('#C5D9F1')
            first_row_format.set_bold()
            first_row_format.set_font_size(11)
            first_row_format.set_text_wrap()
            first_row_format.set_font_name('Calibri')
            worksheet.write_row(row=0, col=0, data=self.header, cell_format=first_row_format)
            if ifNoEntry == True:
                format1 = workbook.add_format()
                format1.set_bold()
                worksheet.set_column('A:A', width=60)
                worksheet.write_string(row=1, col=0, string="Проходы сторудников подразделения отсутствуют.", cell_format=format1)
            workbook.close()
        end = datetime.now()
        print("work time = ", end-start)

        indicator = True
        self.indicator_of_end_work.emit(indicator)

    def readFile(self, file):
        workbook = xlrd.open_workbook(file)
        worksheet = workbook.sheet_by_index(0)
        row = 1
        dataTest = []
        while row != worksheet.nrows:
            for col in range(worksheet.ncols):
                dataTest.append(worksheet.cell_value(row, col))
            row += 1
        return dataTest, worksheet.nrows
