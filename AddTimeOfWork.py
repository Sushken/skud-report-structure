import openpyxl


def AddTimeOfWork(nameOfOutFile, nameOfInFile, value):
    print("vpcp[[apsp[s[asdk[asdl[a")
    wbOut = openpyxl.load_workbook(filename=nameOfOutFile)
    sheetOut = wbOut['Лист1']
    dataOut = sheetOut.values
    dataOut = list(dataOut)

    j = 0
    dataNameOut = []
    while j != sheetOut.max_row:
        dataNameOut.append(dataOut[j][4] + ' ' + dataOut[j][5] + ' ' + dataOut[j][6])
        j += 1
    i = 0
    oneSecondName = 0
    while dataNameOut[i] == dataNameOut[i + 1]:
        oneSecondName += 1
        i += 1

    j = 0
    dataTimeOut = []
    dataTimeOutExit = []
    while j != sheetOut.max_row:
        dataTimeOut.append(dataOut[j][7])
        dataTimeOutExit.append(dataOut[j][8])
        j += 1

    i = 0
    while i < len(dataNameOut):
        if dataTimeOut[i] is not None and dataTimeOutExit[i] is not None and dataNameOut[i] != "Суммарная " \
                                                                                               "продолжительность " \
                                                                                               "рабочего дня за период":
            minus = "=I" + str(i + 2) + "-H" + str(i + 2) + "-K" + str(i + 2)
            sheetOut.cell(row=i + 1, column=10).value = minus
        i += 1
    wbOut.save(filename=nameOfOutFile)

    sheetOut.insert_rows(1, 1)
    sheetOut.cell(row=1, column=1).value = "Компания"
    sheetOut.cell(row=1, column=2).value = "Отдел"
    sheetOut.cell(row=1, column=3).value = "Должность"
    sheetOut.cell(row=1, column=4).value = "Пол"
    sheetOut.cell(row=1, column=5).value = "Фамилия"
    sheetOut.cell(row=1, column=6).value = "Имя"
    sheetOut.cell(row=1, column=7).value = "Отчество"
    sheetOut.cell(row=1, column=8).value = "Первый приход в любой из офисов"
    sheetOut.cell(row=1, column=9).value = "Последний уход из любого офиса"
    sheetOut.cell(row=1, column=10).value = "Продолжительность рабочего дня"
    sheetOut.cell(row=1, column=11).value = "Обеденный перерыв"
    wbOut.save(filename=nameOfOutFile)

    i = 2
    while i <= sheetOut.max_row:
        if sheetOut.cell(row=i, column=9).value is not None:
            if sheetOut.cell(row=i, column=9).value != "Прохода нет":
                if len(str(value)) == 2:
                    if sheetOut.cell(row=i, column=9).value.startswith(str(value)):
                        sheetOut.cell(row=i, column=9).value = ""
                        sheetOut.cell(row=i, column=10).value = ""
                else:
                    if sheetOut.cell(row=i, column=9).value.startswith("0" + str(value)):
                        sheetOut.cell(row=i, column=9).value = ""
                        sheetOut.cell(row=i, column=10).value = ""
            else:
                sheetOut.cell(row=i, column=9).value = ""
                sheetOut.cell(row=i, column=10).value = ""

        i += 1
    wbOut.save(filename=nameOfOutFile)
