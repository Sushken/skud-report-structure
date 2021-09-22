import openpyxl
from PyQt5 import QtCore
import time


class ResolveSeconds(QtCore.QThread):
    percentageChanged = QtCore.pyqtSignal(int)
    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, nameOfOutFile, nameOfInFile, count):
        super().__init__()
        self.nameOfOutFile = nameOfOutFile
        self.nameOfInFile = nameOfInFile
        self.count = count
        self.time = time

    def run(self):
        wbOut = openpyxl.load_workbook(filename=self.nameOfOutFile)
        sheetOut = wbOut['Лист1']
        dataOut = sheetOut.values
        dataOut = list(dataOut)

        wbIn = openpyxl.load_workbook(filename=self.nameOfInFile)
        sheetIn = wbIn['Sheet']
        dataIn = sheetIn.values
        dataIn = list(dataIn)

        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)


        j = 0
        dataTimeOut = []
        dataTimeOutExit = []
        while j != sheetOut.max_row:
            dataTimeOut.append(dataOut[j][7])
            dataTimeOutExit.append(dataOut[j][8])
            j += 1

        j = 0
        dataTimeIn = []
        dataTimeInExit = []
        while j != sheetIn.max_row:
            dataTimeIn.append(dataIn[j][5])
            dataTimeInExit.append(dataIn[j][6])
            j += 1

        for j in range (5):
            self.count += 1
            self.percentageChanged.emit(self.count)

        # Перевод времени входа Ленинградка
        for t in range(0, len(dataTimeOut)):
            if dataTimeOut[t] is not None:
                tmpStr = str(dataTimeOut[t])
                year = tmpStr[:4]
                month = tmpStr[5:7]
                day = tmpStr[8:10]
                time = tmpStr[10:19]
                finalString = day + '.' + month + '.' + year + time
                dataTimeOut[t] = tmpStr
                sheetOut.cell(row=t + 1, column=8).value = finalString

        # Перевод времени выхода Ленинградка
        for t in range(0, len(dataTimeOutExit)):
            if dataTimeOutExit[t] is not None:
                tmpStr = str(dataTimeOutExit[t])
                year = tmpStr[:4]
                month = tmpStr[5:7]
                day = tmpStr[8:10]
                time = tmpStr[10:19]
                finalString = day + '.' + month + '.' + year + time
                dataTimeOutExit[t] = tmpStr
                sheetOut.cell(row=t + 1, column=9).value = finalString
        wbOut.save(filename=self.nameOfOutFile)

        for j in range (5):
            self.count += 1
            self.percentageChanged.emit(self.count)

        # Перевод времени входа Аэродом
        for t in range(0, len(dataTimeIn)):
            if dataTimeIn[t] is not None and len(str(dataTimeIn[t])) == 18:
                tmpStr = str(dataTimeIn[t])
                day = tmpStr[:2]
                month = tmpStr[3:5]
                year = tmpStr[6:10]
                time = tmpStr[11:19]
                finalString = day + '.' + month + '.' + year + ' 0' + time
                dataTimeIn[t] = tmpStr
                sheetIn.cell(row=t + 1, column=6).value = finalString

        # Перевод времени выхода Аэродом
        for t in range(0, len(dataTimeInExit)):
            if dataTimeInExit[t] is not None and len(str(dataTimeInExit[t])) == 18:
                tmpStr = str(dataTimeInExit[t])
                day = tmpStr[:2]
                month = tmpStr[3:5]
                year = tmpStr[6:10]
                time = tmpStr[11:19]
                finalString = day + '.' + month + '.' + year + ' 0' + time
                dataTimeInExit[t] = tmpStr
                sheetIn.cell(row=t + 1, column=7).value = finalString
        wbIn.save(filename=self.nameOfInFile)

        for j in range (5):
            self.count += 1
            self.percentageChanged.emit(self.count)
        self.indicator = True
        self.indicator_of_end_work.emit(self.indicator)
