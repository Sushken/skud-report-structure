import openpyxl
from PyQt5 import QtCore
import time
from openpyxl.styles import Font


class AddSumOfTime(QtCore.QThread):
    percentageChanged = QtCore.pyqtSignal(int)
    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, nameOfOutFile, nameOfInFile, count):
        super().__init__()
        self.nameOfOutFile = nameOfOutFile
        self.nameOfInFile = nameOfInFile
        self.count = count
        self.time = time
        print(self.count, "default count")

    def run(self):
        wbOut = openpyxl.load_workbook(filename=self.nameOfOutFile)
        sheetOut = wbOut['Лист1']
        ws = wbOut.active
        dataOut = sheetOut.values
        dataOut = list(dataOut)

        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)
            # self.time.sleep(0.5)
        print(self.count, "first point")

        j = 0
        dataNameOut = []
        while j != sheetOut.max_row:
            dataNameOut.append(dataOut[j][4] + ' ' + dataOut[j][5] + ' ' + dataOut[j][6])
            j += 1

        for i in range(len(dataNameOut)):
            sheetOut.cell(row=i + 1, column=11).value = "0:45:00"
        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)
            # self.time.sleep(0.5)
        print(self.count, "second point")

        i = 0
        oneSecondName = 0
        while dataNameOut[i] == dataNameOut[i + 1]:
            oneSecondName += 1
            i += 1
        print(oneSecondName)
        print("asd", len(dataNameOut))
        print(int(len(dataNameOut) / (oneSecondName + 1)))
        for i in range(1, int(len(dataNameOut) / (oneSecondName + 1))):
            sheetOut.insert_rows((oneSecondName + 1) * i + i)
        wbOut.save(filename=self.nameOfOutFile)
        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)
            # self.time.sleep(0.5)
        print(self.count, "third point")

        print("after", len(dataNameOut))
        for j in range(1, int(len(dataNameOut) / (oneSecondName + 1)) + 1):
            summ = "=SUM(J" + str((oneSecondName + 1) * (j - 1) + j + 1) + ":J" + str((oneSecondName + 1) * j + j) + ')'
            sheetOut.cell(row=(oneSecondName + 1) * j + j,
                          column=1).value = "Суммарная продолжительность рабочего дня за период"
            ws['A' + str((oneSecondName + 2) * j)].font = Font(bold=True)
            sheetOut.cell(row=(oneSecondName + 1) * j + j, column=2).value = dataOut[(oneSecondName + 1) * j - 1][1]
            sheetOut.cell(row=(oneSecondName + 1) * j + j, column=3).value = dataOut[(oneSecondName + 1) * j - 1][2]
            sheetOut.cell(row=(oneSecondName + 1) * j + j, column=5).value = dataOut[(oneSecondName + 1) * j - 1][4]
            sheetOut.cell(row=(oneSecondName + 1) * j + j, column=6).value = dataOut[(oneSecondName + 1) * j - 1][5]
            sheetOut.cell(row=(oneSecondName + 1) * j + j, column=7).value = dataOut[(oneSecondName + 1) * j - 1][6]
            sheetOut.cell(row=(oneSecondName + 1) * j + j, column=10).value = summ
            ws['J' + str((oneSecondName + 2) * j)].font = Font(bold=True)
        wbOut.save(filename=self.nameOfOutFile)
        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)
            # self.time.sleep(0.5)
        print(self.count, "fourth point")
        self.indicator = True
        self.indicator_of_end_work.emit(self.indicator)



