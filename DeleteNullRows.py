import openpyxl


def main(nameOfFile):
    print("DeleteNullRows")
    wbIn = openpyxl.load_workbook(filename=nameOfFile)
    sheetIn = wbIn['Sheet']
    dataIn = sheetIn.values
    dataIn = list(dataIn)

    i = 0
    dataCompany = []
    while i != sheetIn.max_row:
        dataCompany.append(dataIn[i][0])
        i += 1

    i = 1
    while i <= sheetIn.max_row:
        # print(f'i = {i}\tcell value (i, 1) is {sheetIn.cell(row=i, column=1).value}')
        if sheetIn.cell(row=i, column=1).value is None:
            sheetIn.delete_rows(i, 1)
            # Note the absence of incremental.  Because we deleted a row, we want to stay on the same row because new
            # data will show in the next iteration.
        else:
            i += 1

    wbIn.save(filename=nameOfFile)
