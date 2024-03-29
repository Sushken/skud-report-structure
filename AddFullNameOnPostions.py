# import openpyxl
import openpyxl
from PyQt5 import QtCore
import time


class AddName(QtCore.QThread):
    percentageChanged = QtCore.pyqtSignal(int)
    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, nameOfOutFile, nameOfInFile, count):
        super().__init__()
        self.nameOfOutFile = nameOfOutFile
        self.namaOfInFile = nameOfInFile
        self.count = count

    def run(self):
        wbOut = openpyxl.load_workbook(filename=self.nameOfOutFile)
        sheetOut = wbOut['Лист1']
        dataOut = sheetOut.values
        dataOut = list(dataOut)

        a = list(dataOut[1])

        if (a[7] or a[8]) is None:
            value = "2021.01.01 00:00:01"
            a[7] = value
            a[8] = value
            sheetOut.cell(row=2, column=8).value = a[7]
            sheetOut.cell(row=2, column=9).value = a[8]
            wbOut.save(filename=self.nameOfOutFile)

        sheetOut.delete_rows(1, 1)
        wbOut.save(filename=self.nameOfOutFile)
        sheetOut = wbOut['Лист1']
        dataOut = sheetOut.values
        dataOut = list(dataOut)

        self.count += 2
        self.percentageChanged.emit(self.count)

        i = 0
        dataCompany = []
        dataFamily = []
        dataName = []
        dataMiddleName = []
        dataOtdel = []
        dataWork = []
        while i != sheetOut.max_row:
            if i == sheetOut.max_row / 2:
                self.count += 8
                self.percentageChanged.emit(self.count)
            dataCompany.append(dataOut[i][0])
            dataFamily.append(dataOut[i][4])
            dataName.append(dataOut[i][5])
            dataMiddleName.append(dataOut[i][6])
            dataOtdel.append(dataOut[i][1])
            dataWork.append(dataOut[i][2])
            i += 1

        tmpDataCompany = dataCompany[0]
        tmpDataFamily = dataFamily[0]
        tmpDataName = dataName[0]
        tmpDataMiddleName = dataMiddleName[0]
        tmpDataOtdel = dataOtdel[0]
        tmpDataWork = dataWork[0]
        pre_last_family = 0
        for i in range(1, len(dataFamily)):
            if dataFamily[i] is None:
                sheetOut.cell(row=i + 1, column=1).value = tmpDataCompany
                sheetOut.cell(row=i + 1, column=2).value = tmpDataOtdel
                sheetOut.cell(row=i + 1, column=3).value = tmpDataWork
                sheetOut.cell(row=i + 1, column=5).value = tmpDataFamily
                sheetOut.cell(row=i + 1, column=6).value = tmpDataName
                sheetOut.cell(row=i + 1, column=7).value = tmpDataMiddleName
            else:
                tmpDataCompany = dataCompany[i]
                tmpDataOtdel = dataOtdel[i]
                tmpDataWork = dataWork[i]
                tmpDataFamily = dataFamily[i]
                tmpDataName = dataName[i]
                tmpDataMiddleName = dataMiddleName[i]
                pre_last_family = i
            if i == (len(dataFamily)//3):
                while self.count != 15:
                    self.count += 1
                    self.percentageChanged.emit(self.count)
        wbOut.save(filename=self.nameOfOutFile)

        oneBlock = 0
        k = 1
        while dataName[k] is None:
            oneBlock += 1
            k += 1

        for j in range(1, oneBlock + 1):
            sheetOut.cell(row=pre_last_family + j + 1, column=1).value = tmpDataCompany
            sheetOut.cell(row=pre_last_family + j + 1, column=2).value = tmpDataOtdel
            sheetOut.cell(row=pre_last_family + j + 1, column=3).value = tmpDataWork
            sheetOut.cell(row=pre_last_family + j + 1, column=5).value = tmpDataFamily
            sheetOut.cell(row=pre_last_family + j + 1, column=6).value = tmpDataName
            sheetOut.cell(row=pre_last_family + j + 1, column=7).value = tmpDataMiddleName
        wbOut.save(filename=self.nameOfOutFile)
        while self.count != 20:
            self.count += 1
            self.percentageChanged.emit(self.count)
        self.indicator = True
        self.indicator_of_end_work.emit(self.indicator)
