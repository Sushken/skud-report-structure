import openpyxl
from PyQt5 import QtCore
import time


class AddTimeOfWork(QtCore.QThread):
    percentageChanged = QtCore.pyqtSignal(int)
    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, nameOfOutFile, nameOfInFile, count):
        super().__init__()
        self.nameOfOutFile = nameOfOutFile
        self.nameOfInFile = nameOfInFile
        self.count = count
        self.time = time

    def run(self):
        wbOut = openpyxl.load_workbook(filename=self.nameOfOutFile)
        sheetOut = wbOut['Лист1']
        dataOut = sheetOut.values
        dataOut = list(dataOut)

        j = 0
        dataNameOut = []
        while j != sheetOut.max_row:
            dataNameOut.append(dataOut[j][4] + ' ' + dataOut[j][5] + ' ' + dataOut[j][6])
            j += 1
        i = 0
        oneSecondName = 0
        while dataNameOut[i] == dataNameOut[i + 1]:
            oneSecondName += 1
            i += 1

        for j in range(10):
            self.count += 1
            self.percentageChanged.emit(self.count)

        j = 0
        dataTimeOut = []
        dataTimeOutExit = []
        while j != sheetOut.max_row:
            dataTimeOut.append(dataOut[j][7])
            dataTimeOutExit.append(dataOut[j][8])
            j += 1

        i = 0
        while i < len(dataNameOut):
            if dataTimeOut[i] is not None and dataTimeOutExit[i] is not None and dataNameOut[i] != "Суммарная " \
                                                                                                   "продолжительность " \
                                                                                                   "рабочего дня за период":
                minus = "=I" + str(i + 2) + "-H" + str(i + 2) + "-K" + str(i + 2)
                sheetOut.cell(row=i + 1, column=10).value = minus
            i += 1
        wbOut.save(filename=self.nameOfOutFile)

        for j in range(10):
            self.count += 1
            self.percentageChanged.emit(self.count)

        sheetOut.insert_rows(1, 1)
        sheetOut.cell(row=1, column=1).value = "Компания"
        sheetOut.cell(row=1, column=2).value = "Отдел"
        sheetOut.cell(row=1, column=3).value = "Должность"
        sheetOut.cell(row=1, column=4).value = "Пол"
        sheetOut.cell(row=1, column=5).value = "Фамилия"
        sheetOut.cell(row=1, column=6).value = "Имя"
        sheetOut.cell(row=1, column=7).value = "Отчество"
        sheetOut.cell(row=1, column=8).value = "Первый приход в любой из офисов"
        sheetOut.cell(row=1, column=9).value = "Последний уход из любого офиса"
        sheetOut.cell(row=1, column=10).value = "Продолжительность рабочего дня"
        sheetOut.cell(row=1, column=11).value = "Обеденный перерыв"
        wbOut.save(filename=self.nameOfOutFile)

        self.indicator = True
        self.indicator_of_end_work.emit(self.indicator)
