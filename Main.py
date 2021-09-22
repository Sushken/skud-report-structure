import openpyxl
from PyQt5 import QtCore
import time


class Main(QtCore.QThread):
    percentageChanged = QtCore.pyqtSignal(int)
    indicator_of_end_work = QtCore.pyqtSignal(bool)

    def __init__(self, nameOfOutFile, nameOfInFile, count):
        super().__init__()
        self.nameOfOutFile = nameOfOutFile
        self.nameOfInFile = nameOfInFile
        self.count = count
        self.time = time

    def run(self):
        wbIn = openpyxl.load_workbook(filename=self.nameOfInFile)
        sheetIn = wbIn['Sheet']
        dataIn = sheetIn.values
        dataIn = list(dataIn)

        wbOut = openpyxl.load_workbook(filename=self.nameOfOutFile)
        sheetOut = wbOut['Лист1']
        dataOut = sheetOut.values
        dataOut = list(dataOut)

        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)

        # СОЗДАНИЕ ЛИСТОВ С ФАМИЛИЯМИ И ВРЕМЕНЕМ ПРИХОДА И УХОДА
        i = 0
        dataNameIn = []
        while i != sheetIn.max_row:
            dataNameIn.append(dataIn[i][3])
            i += 1
        j = 0
        dataNameOut = []
        while j != sheetOut.max_row:
            dataNameOut.append(dataOut[j][4] + ' ' + dataOut[j][5] + ' ' + dataOut[j][6])
            j += 1

        i = 0
        oneSecondName = 0
        while dataNameOut[i] == dataNameOut[i + 1]:
            oneSecondName += 1
            i += 1

        i = 0
        dataTimeIn = []
        dataTimeInExit = []
        while i != sheetIn.max_row:
            dataTimeIn.append(dataIn[i][5])
            dataTimeInExit.append(dataIn[i][6])
            i += 1

        j = 0
        dataTimeOut = []
        dataTimeOutExit = []
        while j != sheetOut.max_row:
            dataTimeOut.append(dataOut[j][7])
            dataTimeOutExit.append(dataOut[j][8])
            j += 1

        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)

        print("Сравнение времени входа")
        # Сравнение времени входа
        for j in range(len(dataNameIn)):
            i = 0
            while i <= len(dataNameOut):
                for h in range(oneSecondName):
                    if (i + h) < len(dataNameOut):
                        if dataNameIn[j] == dataNameOut[i + h]:
                            tmpData = str(dataTimeIn[j])[0:2]
                            if tmpData != "Пр":
                                if dataTimeOut[i + int(tmpData) - 1] is None:
                                    sheetOut.cell(row=i + int(tmpData), column=8).value = dataTimeIn[j]
                                    break
                                else:
                                    if (str(dataTimeIn[j])[:10] == str(dataTimeOut[i + int(tmpData) - 1])[:10]) and (
                                            str(dataTimeIn[j])[11:] < str(dataTimeOut[i + int(tmpData) - 1])[11:]):
                                        sheetOut.cell(row=i + int(tmpData), column=8).value = dataTimeIn[j]
                                        break
                            else:
                                if dataTimeOut[i + int(str(dataTimeInExit[j])[0:2]) - 1] is None:
                                    sheetOut.cell(row=i + int(str(dataTimeInExit[j])[0:2]), column=8).value = \
                                    dataTimeIn[j]
                        else:
                            break
                    else:
                        break
                i += oneSecondName + 1
        wbOut.save(filename=self.nameOfOutFile)
        print("Время входа поменялось!")

        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)

        # Сравнение времени выхода
        for j in range(len(dataNameIn)):
            i = 0
            while i <= len(dataNameOut):
                for h in range(oneSecondName):
                    if (i + h) < len(dataNameOut):
                        if dataNameIn[j] == dataNameOut[i + h]:
                            tmpData = str(dataTimeInExit[j])[0:2]
                            if tmpData != "Пр":
                                if dataTimeOutExit[i + int(tmpData) - 1] is None:
                                    sheetOut.cell(row=i + int(tmpData), column=9).value = dataTimeInExit[j]
                                    break
                                else:
                                    if (str(dataTimeInExit[j])[:10] == str(dataTimeOutExit[i + int(tmpData) - 1])[
                                                                       :10]) and (
                                            str(dataTimeInExit[j])[11:] > str(dataTimeOutExit[i + int(tmpData) - 1])[
                                                                          11:]):
                                        sheetOut.cell(row=i + int(tmpData), column=9).value = dataTimeInExit[j]
                                        break
                            else:
                                if (dataTimeOutExit[i + int(str(dataTimeIn[j][0:2])) - 1]) is None:
                                    sheetOut.cell(row=i + int(str(dataTimeIn[j][0:2])), column=9).value = \
                                    dataTimeInExit[j]
                        else:
                            break
                    else:
                        break
                i += oneSecondName + 1
        print("Время выхода поменялось!")
        wbOut.save(filename=self.nameOfOutFile)

        for j in range(5):
            self.count += 1
            self.percentageChanged.emit(self.count)

        self.indicator = True
        self.indicator_of_end_work.emit(self.indicator)
